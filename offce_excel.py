import openpyxl
workbook = openpyxl.Workbook()
worksheet = workbook.create_sheet(title='Sheet1')
for row in range(1, 11):
    for column in range(1, 11):
        cell = worksheet.cell(row=row, column=column)
        cell.value = (row - 1) * 10 + column
workbook.save('sample.xlsx')